from openpyxl import load_workbook
import time


class Product:

    def __init__(self, row):
        self.row = row
        self.table = [cell.value for cell in self.row]
        self.image = self.table[1]
        self.name = self.table[4]
        self.voltage = self.table[6]
        self.watt = self.table[5]
        self.light = self.table[7]
        self.add_info = self.table[8]
        self.opt_price = self.table[10]
        self.roz_price = self.table[9]
        self.category = self.table[3]

    def __repr__(self):
        return f"<Product '{self.name}'>"

    def get_message_text(self):
        return (
            (f"{self.name}\n" if self.name else "")
            + (f"Напряжение (Вольтаж) - {self.voltage}" if self.voltage else "")
            + (f"\nМощность (Ваттность) - {self.watt}" if self.watt else "")
            + (f"\nСвет (Кельвин) - {self.light}" if self.light else "")
            + (f"\nДоп. информация: {self.add_info}" if self.add_info else "")
            + (
                f"\n\n💵Цена (оптовая): от {round(self.opt_price, 2)} тг."
                if self.opt_price
                else ""
            )
            + (
                f"\n💵Цена (розницы): от {round(self.roz_price, 2)} тг."
                if self.roz_price
                else ""
            )
            + (
                f"\n\nПроизводство: Китай\n\nМы являемся производителем, что гарантирует отличную цену с качеством!"
            )
            + (f"\n@novella_electric\n\nКонтакты:\n+77082060860 (Есть Whats App)")
            + (f"\nСайт: novella-electric.kz")
            + (f"\nИнстаграм: instagram.com/novella_electric/")
            + (f"\n\nАдреса магазинов:\nг. Алматы")
            + (f"\n📍Бакорда строй сити 1-й этаж, магазин 139")
            + (f"\n📍Северное кольцо 7, Т/ц Байсат, ряд 15, магазин 15")
            + (f"\n\n#{self.category}" if self.category else "")
        )


class Products:
    # products = []

    def __init__(self, *args):
        if len(args) == 1:
            self.products = args[0]

    def load_products(self, path):
        self.products = list()
        wb = load_workbook(path)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            self.products.append(Product(row))

    def get_products(self):
        return self.products

    def show_products(self):
        print(self.products)


if __name__ == "__main__":
    products = Products()
    products.load_products("/home/troy/Projects/work/novella-bot/app/files/test1.xlsx")
    print(products.get_products())
# wb = load_workbook("/home/troy/Projects/work/novella-bot/app/files/test1.xlsx")
# ws = wb.active


# Access the values of cells in the worksheet
# products = []

# prs = Products(12)
# prs.show_products()
# for cell in row:
#    print(cell.value)
